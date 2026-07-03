"""
Reports API.
Generates PDF and Excel status reports for project tracking and steering reviews.
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from datetime import date

# Exporters
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from backend.database import get_db
from backend.models.project import Project
from backend.services.project_service import get_project_stats
from backend.api.auth import get_current_user

router = APIRouter(prefix="/api/reports", tags=["Reports"])

@router.get("/project/{id}/pdf")
def get_pdf_report(id: int, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """Generates a professional PDF executive report for a project."""
    project = db.query(Project).filter(Project.id == id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    stats = get_project_stats(project)
    latest_pred = project.predictions[0] if project.predictions else None

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    story = []
    styles = getSampleStyleSheet()

    # Title Style
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=24,
        leading=28,
        textColor=colors.HexColor("#0f4c81"),
        alignment=0,
        spaceAfter=15
    )

    # Subtitle Style
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=20
    )

    # Header section
    story.append(Paragraph(f"EXECUTIVE STATUS REPORT: {project.project_name.upper()}", title_style))
    story.append(Paragraph(f"Generated on {date.today().isoformat()} | Manager: {project.project_manager_id or 'System'}", subtitle_style))
    story.append(Spacer(1, 10))

    # Project metadata Table
    meta_data = [
        [Paragraph("<b>Project Code:</b>", styles["Normal"]), Paragraph(project.project_id_code, styles["Normal"]),
         Paragraph("<b>Vessel Class:</b>", styles["Normal"]), Paragraph(project.ship_type or "Unknown", styles["Normal"])],
        [Paragraph("<b>Project Cost:</b>", styles["Normal"]), Paragraph(f"{project.project_cost or 0.0} Crores", styles["Normal"]),
         Paragraph("<b>Overall Progress:</b>", styles["Normal"]), Paragraph(f"{stats['overall_progress']}%", styles["Normal"])],
        [Paragraph("<b>Start Date:</b>", styles["Normal"]), Paragraph(project.start_date.isoformat() if project.start_date else "N/A", styles["Normal"]),
         Paragraph("<b>Expected Completion:</b>", styles["Normal"]), Paragraph(project.expected_end_date.isoformat() if project.expected_end_date else "N/A", styles["Normal"])],
    ]

    t_meta = Table(meta_data, colWidths=[100, 150, 100, 150])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f8fafc")),
        ('PADDING', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 20))

    # ML Predictions section
    story.append(Paragraph("<b>Machine Learning Risk Predictions</b>", styles["Heading2"]))
    if latest_pred:
        pred_data = [
            [Paragraph("<b>Predicted Risk Category:</b>", styles["Normal"]), Paragraph(f"{latest_pred.risk_category} Risk", styles["Normal"])],
            [Paragraph("<b>Predicted Delay Percentage:</b>", styles["Normal"]), Paragraph(f"{latest_pred.delay_percentage}%", styles["Normal"])],
            [Paragraph("<b>Predicted Delay Duration:</b>", styles["Normal"]), Paragraph(f"{latest_pred.delay_months} Months", styles["Normal"])],
            [Paragraph("<b>Prediction Confidence:</b>", styles["Normal"]), Paragraph(f"{latest_pred.confidence}%", styles["Normal"])]
        ]
        t_pred = Table(pred_data, colWidths=[200, 300])
        t_pred.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#eff6ff")),
            ('PADDING', (0,0), (-1,-1), 6),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor("#bfdbfe")),
        ]))
        story.append(t_pred)
    else:
        story.append(Paragraph("No prediction logs recorded yet.", styles["Normal"]))
    story.append(Spacer(1, 20))

    # Activities Table
    story.append(Paragraph("<b>Activity Status Breakdown</b>", styles["Heading2"]))
    act_data = [[
        Paragraph("<b>Seq</b>", styles["Normal"]),
        Paragraph("<b>Activity Name</b>", styles["Normal"]),
        Paragraph("<b>Category</b>", styles["Normal"]),
        Paragraph("<b>Planned End</b>", styles["Normal"]),
        Paragraph("<b>Status</b>", styles["Normal"]),
        Paragraph("<b>Delay (Days)</b>", styles["Normal"])
    ]]

    for act in project.activities[:15]:  # Limit to top 15 for page fitting
        act_data.append([
            Paragraph(str(act.sequence_number), styles["Normal"]),
            Paragraph(act.name, styles["Normal"]),
            Paragraph(act.category, styles["Normal"]),
            Paragraph(act.planned_end_date.isoformat() if act.planned_end_date else "N/A", styles["Normal"]),
            Paragraph(act.current_status.value, styles["Normal"]),
            Paragraph(str(act.current_delay_days), styles["Normal"])
        ])

    t_act = Table(act_data, colWidths=[30, 200, 80, 80, 80, 50])
    t_act.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0f4c81")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('PADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    story.append(t_act)

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    filename = f"Project_Report_{project.project_id_code}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/project/{id}/excel")
def get_excel_report(id: int, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """Generates an Excel status spreadsheet for a project."""
    project = db.query(Project).filter(Project.id == id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    stats = get_project_stats(project)

    # Create openpyxl workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Status Summary"

    # Fonts & Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="0F4C81")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    meta_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    border_side = Side(border_style="thin", color="E2E8F0")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Header Title
    ws["A1"] = f"Ship Acquisition PMIS - Executive Report: {project.project_name}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 30

    # Project metadata
    metadata = [
        ("Project ID", project.project_id_code, "Vessel Class", project.ship_type or "N/A"),
        ("Project Cost (Cr)", project.project_cost or 0.0, "Overall Progress", f"{stats['overall_progress']}%"),
        ("Planned Start", project.start_date, "Expected Completion", project.expected_end_date)
    ]

    for row_idx, row_val in enumerate(metadata, start=3):
        ws.cell(row=row_idx, column=1, value=row_val[0]).font = bold_font
        ws.cell(row=row_idx, column=2, value=row_val[1]).font = normal_font
        ws.cell(row=row_idx, column=3, value=row_val[2]).font = bold_font
        ws.cell(row=row_idx, column=4, value=row_val[3]).font = normal_font
        ws.row_dimensions[row_idx].height = 20
        # Style
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = meta_fill
            cell.border = thin_border

    # Activities Section
    ws["A8"] = "Activities Breakdown"
    ws["A8"].font = Font(name="Calibri", size=14, bold=True, color="334155")
    ws.row_dimensions[8].height = 25

    headers = ["Seq", "Activity Name", "Category", "Planned End", "Actual End", "Status", "Completion %", "Delay (Days)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    ws.row_dimensions[9].height = 22

    for row_idx, act in enumerate(project.activities, start=10):
        ws.cell(row=row_idx, column=1, value=act.sequence_number).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=act.name)
        ws.cell(row=row_idx, column=3, value=act.category)
        ws.cell(row=row_idx, column=4, value=act.planned_end_date.isoformat() if act.planned_end_date else "")
        ws.cell(row=row_idx, column=5, value=act.actual_end_date.isoformat() if act.actual_end_date else "")
        ws.cell(row=row_idx, column=6, value=act.current_status.value).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=7, value=act.completion_percentage).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=8, value=act.current_delay_days).alignment = Alignment(horizontal="right")
        
        ws.row_dimensions[row_idx].height = 18
        for col_idx in range(1, 9):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = normal_font
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Project_Report_{project.project_id_code}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ──────────────────────────────────────────────
# JSON STATE REPORT GENERATION (MICROSERVICE ENDPOINTS)
# ──────────────────────────────────────────────

def create_mock_project_from_json(state: dict):
    from datetime import datetime

    class MockPrediction:
        def __init__(self, risk_category, delay_percentage, delay_months, confidence):
            self.risk_category = risk_category
            self.delay_percentage = delay_percentage
            self.delay_months = delay_months
            self.confidence = confidence

    class MockEnum:
        def __init__(self, val):
            self.value = val

    class MockActivity:
        def __init__(self, act):
            self.sequence_number = act.get("sequence_number", 0)
            self.name = act.get("name", "")
            self.category = act.get("category", "Other")
            self.current_status = MockEnum(act.get("current_status", "NotStarted"))
            self.completion_percentage = act.get("completion_percentage", 0.0)
            self.current_delay_days = act.get("current_delay_days", 0)
            
            p_end = act.get("planned_end_date")
            self.planned_end_date = datetime.strptime(p_end.split('T')[0], "%Y-%m-%d").date() if isinstance(p_end, str) else p_end

            a_start = act.get("actual_start_date")
            self.actual_start_date = datetime.strptime(a_start.split('T')[0], "%Y-%m-%d").date() if isinstance(a_start, str) else None

            a_end = act.get("actual_end_date")
            self.actual_end_date = datetime.strptime(a_end.split('T')[0], "%Y-%m-%d").date() if isinstance(a_end, str) else None

    class MockProject:
        def __init__(self, data):
            self.project_name = data.get("project_name", "")
            self.project_id_code = data.get("project_id_code", "")
            self.ship_name = data.get("ship_name", "")
            self.ship_class = data.get("ship_class", "")
            self.ship_type = data.get("ship_type", "Submarine")
            self.project_cost = data.get("project_cost", 5000.0)
            self.customer = data.get("customer", "Indian Navy")
            self.project_manager_id = data.get("project_manager_name", "Manager")
            
            s_date = data.get("start_date")
            self.start_date = datetime.strptime(s_date.split('T')[0], "%Y-%m-%d").date() if isinstance(s_date, str) else s_date
            
            e_date = data.get("expected_end_date")
            self.expected_end_date = datetime.strptime(e_date.split('T')[0], "%Y-%m-%d").date() if isinstance(e_date, str) else e_date
            
            self.activities = [MockActivity(a) for a in data.get("activities", [])]
            
            # Predict fields
            self.predictions = []
            if "predicted_risk_category" in data or "predictedDelayPct" in data:
                self.predictions.append(MockPrediction(
                    data.get("predicted_risk_category", data.get("predictedRiskCategory", "Low")),
                    data.get("predicted_delay_pct", data.get("predictedDelayPct", 0.0)),
                    data.get("predicted_delay_months", data.get("predictedDelayMonths", 0.0)),
                    data.get("prediction_confidence", data.get("predictionConfidence", 85.0))
                ))

    return MockProject(state)


def compute_mock_stats(state: dict):
    acts = state.get("activities", [])
    total = len(acts)
    if total == 0:
        return {
            "overall_progress": 0.0,
            "total_activities": 0,
            "completed_activities": 0,
            "delayed_activities": 0
        }
    completed = sum(1 for a in acts if a.get("current_status") == "Completed")
    delayed = sum(1 for a in acts if a.get("current_status") in ["Delayed", "Blocked"])
    progress = sum(a.get("completion_percentage", 0.0) for a in acts)
    return {
        "overall_progress": round(progress / total, 1),
        "total_activities": total,
        "completed_activities": completed,
        "delayed_activities": delayed
    }


@router.post("/generate/pdf")
def post_pdf_report(state: dict):
    """Generates PDF report from serialized JSON project state."""
    project = create_mock_project_from_json(state)
    stats = compute_mock_stats(state)
    latest_pred = project.predictions[0] if project.predictions else None

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    story = []
    styles = getSampleStyleSheet()

    # Title Style
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=24,
        leading=28,
        textColor=colors.HexColor("#0f4c81"),
        alignment=0,
        spaceAfter=15
    )

    # Subtitle Style
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=20
    )

    # Header section
    story.append(Paragraph(f"EXECUTIVE STATUS REPORT: {project.project_name.upper()}", title_style))
    story.append(Paragraph(f"Generated on {date.today().isoformat()} | Manager: {project.project_manager_id}", subtitle_style))
    story.append(Spacer(1, 10))

    # Project metadata Table
    meta_data = [
        [Paragraph("<b>Project Code:</b>", styles["Normal"]), Paragraph(project.project_id_code, styles["Normal"]),
         Paragraph("<b>Vessel Class:</b>", styles["Normal"]), Paragraph(project.ship_type or "Unknown", styles["Normal"])],
        [Paragraph("<b>Project Cost:</b>", styles["Normal"]), Paragraph(f"{project.project_cost or 0.0} Crores", styles["Normal"]),
         Paragraph("<b>Overall Progress:</b>", styles["Normal"]), Paragraph(f"{stats['overall_progress']}%", styles["Normal"])],
        [Paragraph("<b>Start Date:</b>", styles["Normal"]), Paragraph(project.start_date.isoformat() if project.start_date else "N/A", styles["Normal"]),
         Paragraph("<b>Expected Completion:</b>", styles["Normal"]), Paragraph(project.expected_end_date.isoformat() if project.expected_end_date else "N/A", styles["Normal"])],
    ]

    t_meta = Table(meta_data, colWidths=[100, 150, 100, 150])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f8fafc")),
        ('PADDING', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 20))

    # ML Predictions section
    story.append(Paragraph("<b>Machine Learning Risk Predictions</b>", styles["Heading2"]))
    if latest_pred:
        pred_data = [
            [Paragraph("<b>Predicted Risk Category:</b>", styles["Normal"]), Paragraph(f"{latest_pred.risk_category} Risk", styles["Normal"])],
            [Paragraph("<b>Predicted Delay Percentage:</b>", styles["Normal"]), Paragraph(f"{latest_pred.delay_percentage}%", styles["Normal"])],
            [Paragraph("<b>Predicted Delay Duration:</b>", styles["Normal"]), Paragraph(f"{latest_pred.delay_months} Months", styles["Normal"])],
            [Paragraph("<b>Prediction Confidence:</b>", styles["Normal"]), Paragraph(f"{latest_pred.confidence}%", styles["Normal"])]
        ]
        t_pred = Table(pred_data, colWidths=[200, 300])
        t_pred.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#eff6ff")),
            ('PADDING', (0,0), (-1,-1), 6),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor("#bfdbfe")),
        ]))
        story.append(t_pred)
    else:
        story.append(Paragraph("No prediction logs recorded yet.", styles["Normal"]))
    story.append(Spacer(1, 20))

    # Activities Table
    story.append(Paragraph("<b>Activity Status Breakdown</b>", styles["Heading2"]))
    act_data = [[
        Paragraph("<b>Seq</b>", styles["Normal"]),
        Paragraph("<b>Activity Name</b>", styles["Normal"]),
        Paragraph("<b>Category</b>", styles["Normal"]),
        Paragraph("<b>Planned End</b>", styles["Normal"]),
        Paragraph("<b>Status</b>", styles["Normal"]),
        Paragraph("<b>Delay (Days)</b>", styles["Normal"])
    ]]

    for act in project.activities[:25]:  # Allow more for full overview
        act_data.append([
            Paragraph(str(act.sequence_number), styles["Normal"]),
            Paragraph(act.name, styles["Normal"]),
            Paragraph(act.category, styles["Normal"]),
            Paragraph(act.planned_end_date.isoformat() if act.planned_end_date else "N/A", styles["Normal"]),
            Paragraph(act.current_status.value, styles["Normal"]),
            Paragraph(str(act.current_delay_days), styles["Normal"])
        ])

    t_act = Table(act_data, colWidths=[30, 200, 80, 80, 80, 50])
    t_act.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0f4c81")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('PADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    story.append(t_act)

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    filename = f"Project_Report_{project.project_id_code}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/generate/excel")
def post_excel_report(state: dict):
    """Generates Excel spreadsheet report from serialized JSON project state."""
    project = create_mock_project_from_json(state)
    stats = compute_mock_stats(state)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Status Summary"

    title_font = Font(name="Calibri", size=16, bold=True, color="0F4C81")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    meta_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    border_side = Side(border_style="thin", color="E2E8F0")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    ws["A1"] = f"Ship Acquisition PMIS - Executive Report: {project.project_name}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 30

    metadata = [
        ("Project ID", project.project_id_code, "Vessel Class", project.ship_type or "N/A"),
        ("Project Cost (Cr)", project.project_cost or 0.0, "Overall Progress", f"{stats['overall_progress']}%"),
        ("Planned Start", project.start_date, "Expected Completion", project.expected_end_date)
    ]

    for row_idx, row_val in enumerate(metadata, start=3):
        ws.cell(row=row_idx, column=1, value=row_val[0]).font = bold_font
        ws.cell(row=row_idx, column=2, value=row_val[1]).font = normal_font
        ws.cell(row=row_idx, column=3, value=row_val[2]).font = bold_font
        ws.cell(row=row_idx, column=4, value=row_val[3]).font = normal_font
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = meta_fill
            cell.border = thin_border

    ws["A8"] = "Activities Breakdown"
    ws["A8"].font = Font(name="Calibri", size=14, bold=True, color="334155")
    ws.row_dimensions[8].height = 25

    headers = ["Seq", "Activity Name", "Category", "Planned End", "Actual End", "Status", "Completion %", "Delay (Days)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    ws.row_dimensions[9].height = 22

    for row_idx, act in enumerate(project.activities, start=10):
        ws.cell(row=row_idx, column=1, value=act.sequence_number).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=act.name)
        ws.cell(row=row_idx, column=3, value=act.category)
        ws.cell(row=row_idx, column=4, value=act.planned_end_date.isoformat() if act.planned_end_date else "")
        ws.cell(row=row_idx, column=5, value=act.actual_end_date.isoformat() if act.actual_end_date else "")
        ws.cell(row=row_idx, column=6, value=act.current_status.value).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=7, value=act.completion_percentage).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=8, value=act.current_delay_days).alignment = Alignment(horizontal="right")
        
        ws.row_dimensions[row_idx].height = 18
        for col_idx in range(1, 9):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = normal_font
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Project_Report_{project.project_id_code}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
